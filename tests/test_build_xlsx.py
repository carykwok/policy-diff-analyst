from openpyxl import load_workbook
from scripts.models import DiffReport, DiffItem, StrengthScore
from scripts.build_xlsx import build_xlsx

def _sample_report():
    return DiffReport(
        old_doc_title="2023",
        new_doc_title="2024",
        items=[
            DiffItem(layer="A1", change_type="modified", old="GDP 5.5%", new="GDP 5%", note="目标下调"),
            DiffItem(layer="A3", change_type="added", old="", new="新质生产力", note=""),
        ],
        strength=[StrengthScore(f"A{i}_x", 2.0, 3.0) for i in range(1, 7)],
        term_freq={"新质生产力": {"old": 0, "new": 12}, "房地产": {"old": 5, "new": 3}},
    )

def test_xlsx_has_four_sheets(tmp_path):
    out = tmp_path / "data.xlsx"
    build_xlsx(_sample_report(), out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["指标对比", "词频统计", "政策强度", "差异清单"]

def test_diff_list_sheet_rows_match_report(tmp_path):
    report = _sample_report()
    out = tmp_path / "data.xlsx"
    build_xlsx(report, out)
    wb = load_workbook(out)
    ws = wb["差异清单"]
    # header + 2 items
    assert ws.max_row == len(report.items) + 1
