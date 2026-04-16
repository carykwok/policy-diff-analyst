from pathlib import Path
from openpyxl import Workbook
from scripts.models import DiffReport
from scripts.score_model import top_n_term_freq, strength_to_dataframe

A1_QUANTITATIVE_KEYS = ("GDP", "国内生产总值", "赤字率", "CPI", "居民消费价格", "城镇新增就业")

def _write_df(ws, df) -> None:
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))

def build_xlsx(report: DiffReport, out: Path) -> None:
    wb = Workbook()

    # Sheet 1: 指标对比 — rows from term_freq whose term is a quantitative key
    ws1 = wb.active
    ws1.title = "指标对比"
    ws1.append(["指标", f"旧版({report.old_doc_title})", f"新版({report.new_doc_title})", "差值"])
    for term, freq in report.term_freq.items():
        if any(k in term for k in A1_QUANTITATIVE_KEYS):
            ws1.append([term, freq["old"], freq["new"], freq["new"] - freq["old"]])

    # Sheet 2: 词频统计 — top-50
    ws2 = wb.create_sheet("词频统计")
    _write_df(ws2, top_n_term_freq(report.term_freq, n=50))

    # Sheet 3: 政策强度
    ws3 = wb.create_sheet("政策强度")
    _write_df(ws3, strength_to_dataframe(report.strength))

    # Sheet 4: 差异清单
    ws4 = wb.create_sheet("差异清单")
    ws4.append(["层级", "变化类型", "旧版表述", "新版表述", "备注"])
    for item in report.items:
        ws4.append([item.layer, item.change_type, item.old, item.new, item.note])

    wb.save(out)
